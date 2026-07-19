import os
import io
import csv
import uuid
import json
import time
import threading

from flask import Flask, render_template, request, jsonify, Response, stream_with_context
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

import scanner

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB upload limit

_DEFAULT_KEYS = [k.strip() for k in os.getenv("VT_API_KEYS", "").split(",") if k.strip()]
_MAX_WORKERS = int(os.getenv("MAX_WORKERS", 10))

# Pre-initialize key states so they show in UI before first scan
scanner.init_key_states(_DEFAULT_KEYS)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/scan", methods=["POST"])
def start_scan():
    # Collect IPs from textarea or uploaded file
    raw_text = request.form.get("ips", "")

    uploaded = request.files.get("file")
    if uploaded and uploaded.filename:
        content = uploaded.read().decode("utf-8", errors="ignore")
        raw_text = raw_text + "\n" + content

    # Also accept JSON body
    if not raw_text and request.is_json:
        body = request.get_json(silent=True) or {}
        raw_text = body.get("ips", "")

    ips = scanner.parse_ips(raw_text)
    if not ips:
        return jsonify({"error": "No valid IPs found"}), 400

    # API keys: from form, env, or JSON
    keys_raw = request.form.get("api_keys", "") or (request.get_json(silent=True) or {}).get("api_keys", "")
    if keys_raw:
        keys = [k.strip() for k in keys_raw.replace("\n", ",").split(",") if k.strip()]
    else:
        keys = _DEFAULT_KEYS

    if not keys:
        return jsonify({"error": "No API keys provided"}), 400

    workers = int(request.form.get("workers", _MAX_WORKERS))
    workers = max(1, min(workers, 50))

    job_id = str(uuid.uuid4())
    scanner.start_scan(job_id, ips, keys, workers)

    return jsonify({"job_id": job_id, "total": len(ips)})


@app.route("/progress/<job_id>")
def progress(job_id):
    def generate():
        while True:
            job = scanner.get_job(job_id)
            if not job:
                yield f"data: {json.dumps({'error': 'Job not found'})}\n\n"
                break
            payload = {
                "status": job["status"],
                "done": job["done"],
                "total": job["total"],
                "pct": round(job["done"] / job["total"] * 100, 1) if job["total"] else 0,
                "keys_exhausted": job.get("keys_exhausted", False),
                "key_status": scanner.get_key_status(),
            }
            yield f"data: {json.dumps(payload)}\n\n"
            if job["status"] in ("complete", "cancelled", "error"):
                break
            time.sleep(0.5)

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.route("/results/<job_id>")
def results(job_id):
    job = scanner.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "total": job["total"],
        "done": job["done"],
        "results": job["results"],
        "started_at": job["started_at"],
        "finished_at": job["finished_at"],
    })


@app.route("/key-status")
def key_status():
    return jsonify(scanner.get_key_status())


@app.route("/cancel/<job_id>", methods=["POST"])
def cancel(job_id):
    ok = scanner.cancel_job(job_id)
    return jsonify({"cancelled": ok})


@app.route("/download/<job_id>/<fmt>")
def download(job_id, fmt):
    job = scanner.get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    results = job["results"]
    headers_row = ["IP", "Verdict", "Malicious", "Suspicious", "Harmless", "Undetected", "Country", "Owner/ASN", "ASN", "Error"]

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers_row)
        for r in results:
            writer.writerow([
                r["ip"], r["verdict"], r["malicious"], r["suspicious"],
                r["harmless"], r["undetected"], r["country"],
                r["owner"], r["asn"], r.get("error", "")
            ])
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=scan_{job_id[:8]}.csv"},
        )

    elif fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scan Results"

        # Header style
        header_fill = PatternFill("solid", fgColor="1a1a2e")
        header_font = Font(color="00d4ff", bold=True)
        ws.append(headers_row)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        fill_mal = PatternFill("solid", fgColor="3d0000")
        fill_sus = PatternFill("solid", fgColor="3d2800")
        fill_clean = PatternFill("solid", fgColor="003d00")
        fill_err = PatternFill("solid", fgColor="1a1a1a")

        for r in results:
            row = [
                r["ip"], r["verdict"], r["malicious"], r["suspicious"],
                r["harmless"], r["undetected"], r["country"],
                r["owner"], r["asn"], r.get("error", "")
            ]
            ws.append(row)
            verdict = r["verdict"]
            row_idx = ws.max_row
            fill = (fill_mal if verdict == "malicious" else
                    fill_sus if verdict == "suspicious" else
                    fill_err if verdict == "error" else fill_clean)
            for cell in ws[row_idx]:
                cell.fill = fill

        # Column widths
        widths = [18, 12, 11, 11, 11, 12, 10, 30, 10, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=scan_{job_id[:8]}.xlsx"},
        )

    return jsonify({"error": "Unknown format"}), 400


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    debug = os.getenv("FLASK_ENV", "production") == "development"
    app.run(host="0.0.0.0", port=port, debug=debug, threaded=True)
